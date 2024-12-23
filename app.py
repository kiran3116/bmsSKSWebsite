   
# new code 
import os
import sqlite3
from flask import Flask, jsonify, request, render_template, redirect, url_for, session, flash
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from twilio.rest import Client
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient
import pandas as pd
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import google.auth
from googleapiclient.discovery import build
from io import BytesIO


# Load environment variables
load_dotenv()
SECRET_KEY = os.getenv('SECRET_KEY')
TWILIO_ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID')
TWILIO_AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN')
TWILIO_PHONE_NUMBER = os.getenv('TWILIO_PHONE_NUMBER')

# Initialize Flask app
app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

# Twilio client initialization
twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# File upload configuration
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# SQLite database initialization
def init_sqlite_db():
    with sqlite3.connect("users.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT NOT NULL,
                password TEXT NOT NULL
            )
        """)
        conn.commit()

init_sqlite_db()

# Helper Functions
def is_logged_in():
    """Check if the user is logged in."""
    return 'username' in session

def send_email(subject, body, recipient_email):
    """Send an email using SMTP."""
    sender_email = "your_email@example.com"
    sender_password = "your_email_password"
    
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = recipient_email
    message['Subject'] = subject
    message.attach(MIMEText(body, 'plain'))
    
    try:
        with smtplib.SMTP('smtp.example.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_email, message.as_string())
    except Exception as e:
        print(f"Email error: {e}")

# Flask Routes

## User Authentication Routes
@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration."""
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        hashed_password = generate_password_hash(password)
        
        try:
            with sqlite3.connect("users.db") as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT INTO users (username, email, password) VALUES (?, ?, ?)",
                               (username, email, hashed_password))
                conn.commit()
                flash('Registration successful! Please log in.', 'success')
                return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Please choose another.', 'error')
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login."""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        with sqlite3.connect("users.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()
        
        if user and check_password_hash(user[3], password):
            session['username'] = username
            return redirect(url_for('home'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """User logout."""
    session.pop('username', None)
    return redirect(url_for('login'))

## Main Application Routes
@app.route('/')
def home():
    """Home page."""
    if is_logged_in():
        return render_template('home.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    """Dashboard page."""
    if is_logged_in():
        return render_template('dashboard.html', username=session['username'])
    return redirect(url_for('login'))


@app.route('/normal_sms')
def normal_sms():
    if is_logged_in():
        return render_template('index.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/excel')
def excel():
    if is_logged_in():
        return render_template('excel.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/advanced')
def advanced_sms():
    if is_logged_in():
        return render_template('advanced.html', username=session['username'])
    return redirect(url_for('login'))



# for mongo db
@app.route('/google_sms')
def normal_google():
    if is_logged_in():
        return render_template('googlesheet.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/mongo_column')
def mongo_column():
    if is_logged_in():
        return render_template('excel.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/mongo_advanced')
def mongo_advanced():
    if is_logged_in():
        return render_template('advanced.html', username=session['username'])
    return redirect(url_for('login'))


# end

# start of google sheet




@app.route('/fetch-sheet', methods=['POST'])
def fetch_google_sheet():
    try:
        sheet_url = request.json.get('sheet_url')
        if not sheet_url:
            return jsonify({'error': 'Google Sheet URL is required'}), 400

        # Extract spreadsheet ID from URL
        spreadsheet_id = sheet_url.split("/d/")[1].split("/")[0]

        # Setup Google Sheets API
        creds, _ = google.auth.default()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Fetch sheet details
        sheet_metadata = sheet.get(spreadsheetId=spreadsheet_id).execute()
        sheets = sheet_metadata.get('sheets', [])
        sheet_names = [s['properties']['title'] for s in sheets]

        return jsonify({'sheet_names': sheet_names})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/fetch-data', methods=['POST'])
def fetch_sheet_data():
    try:
        sheet_url = request.json.get('sheet_url')
        selected_sheet = request.json.get('sheet_name')
        if not sheet_url or not selected_sheet:
            return jsonify({'error': 'Sheet URL and Sheet Name are required'}), 400

        spreadsheet_id = sheet_url.split("/d/")[1].split("/")[0]

        creds, _ = google.auth.default()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Fetch selected sheet data
        result = sheet.values().get(spreadsheetId=spreadsheet_id, range=selected_sheet).execute()
        values = result.get('values', [])

        return jsonify({'data': values})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/google_send-sms', methods=['POST'])
def google_send_sms():
    try:
        data = request.json
        message_text = data.get('message')
        recipients = data.get('recipients')

        if not message_text or not recipients:
            return jsonify({'error': 'Message text and recipients are required'}), 400

        # Initialize Twilio client
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

        # Send messages to recipients
        response_details = []
        for recipient in recipients:
            try:
                message = client.messages.create(
                    body=message_text,
                    from_=TWILIO_PHONE_NUMBER,
                    to=recipient
                )
                response_details.append({'recipient': recipient, 'status': 'sent', 'sid': message.sid})
            except Exception as e:
                response_details.append({'recipient': recipient, 'status': 'failed', 'error': str(e)})

        return jsonify({'results': response_details})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# end


#  for maria db
@app.route('/normal_maria')
def normal_maria():
    if is_logged_in():
        return render_template('index.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/maria_column')
def maria_column():
    if is_logged_in():
        return render_template('excel.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/maria_advanced')
def maria_advanced():
    if is_logged_in():
        return render_template('advanced.html', username=session['username'])
    return redirect(url_for('login'))


# end


@app.route('/contacts')
def contacts():
    """Contact us page."""
    return render_template('contactus.html')

@app.route('/profile')
def profile():
    """User profile page."""
    if is_logged_in():
        return render_template('profile.html', username=session['username'])
    return redirect(url_for('login'))

@app.route('/settings')
def settings():
    """User settings page."""
    if is_logged_in():
        return render_template('settings.html', username=session['username'])
    return redirect(url_for('login'))

## MongoDB Routes
# Initialize MongoDB
mongo_client = MongoClient('mongodb://localhost:27017/')
mongo_db = mongo_client['your_database_name']

# Initialize Twilio Client
# twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# ---------------------------- MongoDB Routes -------------------------------- #

@app.route('/collections', methods=['GET'])
def get_collections():
    """List all MongoDB collections."""
    collections = mongo_db.list_collection_names()
    return jsonify({'collections': collections})


@app.route('/fields', methods=['GET'])
def get_fields():
    """Get fields (columns) from a MongoDB collection."""
    collection_name = request.args.get('collection')
    if not collection_name:
        return jsonify({'fields': []})
    collection = mongo_db[collection_name]
    sample_doc = collection.find_one()
    fields = list(sample_doc.keys()) if sample_doc else []
    return jsonify({'fields': fields})


@app.route('/mongo_send_sms', methods=['POST'])
def mongo_send_sms():
    """Send SMS to phone numbers fetched from MongoDB."""
    data = request.json
    collection_name = data['collection']
    field_name = data['field']
    message = data['message']
    
    collection = mongo_db[collection_name]
    recipients = collection.find({}, {field_name: 1, '_id': 0})
    success_count = 0
    
    for recipient in recipients:
        phone_number = recipient.get(field_name)
        if phone_number:
            try:
                twilio_client.messages.create(
                    body=message,
                    from_=TWILIO_PHONE_NUMBER,
                    to=phone_number
                )
                success_count += 1
            except Exception as e:
                print(f"SMS error for {phone_number}: {e}")
    
    return jsonify({'message': f'Successfully sent {success_count} messages.'})

# ---------------------------- Excel Upload Routes ---------------------------- #

@app.route('/upload', methods=['POST'])
def upload_file():
    """Upload Excel file and return sheet names."""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    
    try:
        workbook = load_workbook(filepath)
        sheets = workbook.sheetnames
        return jsonify({"sheets": sheets})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/columns', methods=['POST'])
def get_columns():
    """Get column names from a selected Excel sheet."""
    data = request.get_json()
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], data['filename'])
    sheet_name = data['sheet']
    
    try:
        workbook = load_workbook(filepath)
        sheet = workbook[sheet_name]
        columns = [cell.value for cell in sheet[1]]
        return jsonify({"columns": columns})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/send_sms', methods=['POST'])
def send_sms():
    """Send SMS messages to numbers extracted from an Excel file."""
    data = request.get_json()
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], data['filename'])
    sheet_name = data['sheet']
    message_text = data['message']
    column_index = int(data['column'])
    
    try:
        workbook = load_workbook(filepath)
        sheet = workbook[sheet_name]
        phone_numbers = [
            str(row[column_index].value)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row)
            if row[column_index].value
        ]
        
        for number in phone_numbers:
            twilio_client.messages.create(
                body=message_text,
                from_=TWILIO_PHONE_NUMBER,
                to=number
            )
        
        return jsonify({"success": True, "message": f"SMS sent to {len(phone_numbers)} contacts."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/excel_send_sms', methods=['POST'])
def excel_send_sms():
    try:
        data = request.get_json()
        phones = data.get('phones', [])
        messages = data.get('messages', [])

        if not phones or not messages:
            return jsonify({"error": "Phone numbers or messages are missing."}), 400

        sent_count = 0
        for phone, message in zip(phones, messages):
            if phone and message:
                twilio_client.messages.create(
                    body=message,
                    from_=TWILIO_PHONE_NUMBER,
                    to=phone
                )
                sent_count += 1

        return jsonify({"message": f"SMS sent to {sent_count} recipients."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/advanced_send_sms', methods=['POST'])
def advanced_send_sms():
    """Send SMS using a message template and mapped Excel data."""
    data = request.json
    template_message = data.get("templateMessage")
    phones = data.get('phones', [])
    placeholders = data.get('placeholders', [])

    if not phones or not template_message:
        return jsonify({"error": "Missing phone numbers or message template."}), 400

    sent_count = 0
    try:
        for phone, placeholder_values in zip(phones, placeholders):
            formatted_message = template_message.format(**placeholder_values)
            twilio_client.messages.create(
                body=formatted_message,
                from_=TWILIO_PHONE_NUMBER,
                to=phone
            )
            sent_count += 1
        
        return jsonify({"message": f"Successfully sent SMS to {sent_count} recipients."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ---------------------------- Contact Form Route ---------------------------- #

@app.route('/send_contact_form', methods=['POST'])
def send_contact_form():
    """Send a contact form submission via email."""
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        subject = request.form['subject']
        message = request.form['message']

        email_content = f"Name: {name}\nEmail: {email}\nSubject: {subject}\nMessage:\n{message}"

        try:
            # Replace this with your email logic
            print("Email Sent:")
            print(email_content)
            flash("Your message has been sent successfully!", "success")
        except Exception as e:
            flash(f"An error occurred: {e}", "error")

        return redirect(url_for('contact_us'))


@app.route('/contacts')
def contact_us():
    """Render the contact us page."""
    return render_template('contactus.html')

# ---------------------------- Application Runner ---------------------------- #

if __name__ == '__main__':
    # Ensure upload folder exists
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(host='127.0.0.1', port=5000, debug=True)
